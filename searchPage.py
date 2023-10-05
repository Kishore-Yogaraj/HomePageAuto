# Import necessary modules from the selenium library
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from openpyxl import Workbook
import pyautogui
import time

def returnToHome():
    time.sleep(2)
    logo = driver.find_element(by=By.CSS_SELECTOR, value='[alt="banner"]')
    if logo.is_displayed():
       print("Logo return to home is displayed")
    else:
        print("Logo return to home is not displayed")
    logo.click()
    time.sleep(2)

def handle_element (element, description):
    displayed_str = f"{description} is displayed"
    not_displayed_str = f"{description} is not displayed"
    enabled_str = f"{description} is enabled"
    disabled_str = f"{description} is disabled"

    if element.is_displayed():
        print(displayed_str)
        buttonDisplayed.append(displayed_str)
    else:
        print(not_displayed_str)
        buttonNotDisplayed.append(not_displayed_str)

    if element.is_enabled():
        print(enabled_str)
        buttonEnabled.append(enabled_str)
    else:
        print(disabled_str)
        buttonDisabled.append(disabled_str)

def click_and_record (element, description, action_list):
    element.click()
    success_message = f"{description} Passed"
    print(success_message)
    action_list.append(success_message)

wb = Workbook()

ws = wb.active

clickActionPassed = []
hoverActionPassed = []
buttonDisplayed = []
buttonEnabled = []
buttonNotDisplayed = []
buttonDisabled = []
hoverActionFailed = []
clickActionFailed = []

# Create an Options object to store Chrome options
options = Options()

# Window size
options.add_argument("--window-size=1920,1080")

# This allows the Chrome window to remain open after the script finishes
options.add_experimental_option("detach", True)

# options.add_argument("--headless=new")

# Initialize the Chrome driverai-user   vq4alm56

driver = webdriver.Chrome(options=options)

# Access link to website
url = "https://wralstaging.eonmedia.ai/"
driver.get(url)

# Wait for pop up to open
time.sleep(2)

# Type the username and press TAB to switch to the password field
pyautogui.typewrite('ai-user')
pyautogui.press('tab')

# Type the password and press Enter to submit
pyautogui.typewrite('vq4alm56')

pyautogui.press('enter')

time.sleep(3)

login = driver.find_element(by=By.ID, value="log-in-button")

if login.is_displayed():
    logInButtonDisplayed = "Login is displayed"
    print(logInButtonDisplayed)
    buttonDisplayed.append(logInButtonDisplayed)
else:
    logInButtonNotDisplayed = "Login is not displayed"
    print(logInButtonNotDisplayed)
    buttonNotDisplayed.append(logInButtonNotDisplayed)

if login.is_enabled():
    logInButtonEnabled = "Log In button is enabled"
    print(logInButtonEnabled)
    buttonEnabled.append(logInButtonEnabled)
else:
    logInButtonDisabled = "Log In button is disabled"
    print(logInButtonDisabled)
    buttonDisabled.append(logInButtonDisabled)

login.click()
logInPass = "Log In Passed"
print(logInPass)
clickActionPassed.append(logInPass)

time.sleep(3)

emailInput = driver.find_element(by=By.XPATH, value="/html/body/div[1]/div/div[1]/div[6]/div/div[1]/form/input")
if emailInput.is_displayed():
    emailInputDisplayed = "Email input text box is displayed"
    print(emailInputDisplayed)
    buttonDisplayed.append((emailInputDisplayed))
else:
    emailInputNotDisplayed = "Email input text box is not displayed"
    print(emailInputNotDisplayed)
    buttonNotDisplayed.append(emailInputNotDisplayed)

emailInput.click()
emailInput.send_keys("kishore.yogaraj@gmail.com")
emailInput.send_keys(Keys.TAB)
emailInputClickAction = "Username passed"
print(emailInputClickAction)
clickActionPassed.append(emailInputClickAction)

time.sleep(1)

passwordInput = driver.find_element(by=By.XPATH, value="/html/body/div[1]/div/div[1]/div[6]/div/div[1]/form/div[3]/input")
if passwordInput.is_displayed():
    passwordInputDisplayed = "Password input text box is displayed"
    print(passwordInputDisplayed)
    buttonDisplayed.append(passwordInputDisplayed)
else:
    passwordInputNotDisplayed = "Password input text box is not displayed"
    print(passwordInputNotDisplayed)
    buttonNotDisplayed.append(passwordInputNotDisplayed)

passwordInput.click()
passwordInput.send_keys("password1")
passwordInputClickAction = "Password passed"
print(passwordInputClickAction)
clickActionPassed.append(passwordInputClickAction)

time.sleep(1)

hidePassword = driver.find_element(by=By.XPATH, value="/html/body/div[1]/div/div[1]/div[6]/div/div[1]/form/div[3]/div")
if hidePassword.is_displayed():
    hidePasswordDisplayed = "Hide password button is displayed"
    print(hidePasswordDisplayed)
    buttonDisplayed.append(hidePasswordDisplayed)
else:
    hidePasswordNotDisplayed = "Hide button password is not displayed"
    print(hidePasswordNotDisplayed)
    buttonNotDisplayed.append(hidePasswordNotDisplayed)

if hidePassword.is_enabled():
    hidePasswordEnabled = "Hide password button is enabled"
    print(hidePasswordEnabled)
    buttonEnabled.append(hidePasswordEnabled)
else:
    hidePasswordDisabled = "Hide password button is disabled"
    print(hidePasswordDisabled)
    buttonEnabled.append(hidePasswordDisabled)

hidePassword.click()

viewPasswordClickAction = "View password successful"
print(viewPasswordClickAction)
clickActionPassed.append(viewPasswordClickAction)
time.sleep(2)

hidePassword.click()

hidePasswordClickAction = "Hide password successful"
print(hidePasswordClickAction)
clickActionPassed.append(hidePasswordClickAction)

time.sleep(1)

signIn = driver.find_element(by=By.XPATH, value="/html/body/div[1]/div/div[1]/div[6]/div/div[1]/form/div[4]/button")
if signIn.is_displayed():
    signInDisplayed = "Sign in button is displayed"
    print(signInDisplayed)
    buttonDisplayed.append(signInDisplayed)
else:
    signInNotDisplayed = "Sign in button is not displayed"
    print(signInNotDisplayed)
    buttonNotDisplayed.append(signInNotDisplayed)

if signIn.is_enabled():
    signInEnabled = "Sign In button is enabled"
    print(signInEnabled)
    buttonEnabled.append(signInEnabled)
else:
    signInDisabled = "Sign In button is disabled"
    print(signInDisabled)
    buttonDisabled.append(signInDisabled)

signIn.click()
signInClickAction = "Sign in button passed"
print(signInClickAction)
clickActionPassed.append(signInClickAction)

loginProcessClickAction = "Login process passed"
print(loginProcessClickAction)
clickActionPassed.append(loginProcessClickAction)

searchBar = driver.find_element(by=By.CLASS_NAME, value="elastic-search-input")
if searchBar.is_displayed():
    searchBarDisplayed = "Search Bar is displayed"
    print(searchBarDisplayed)
    buttonDisplayed.append(searchBarDisplayed)
else:
    searchBarNotDisplayed = "Search Bar is not displayed"
    print(searchBarNotDisplayed)
    buttonNotDisplayed.append(searchBarNotDisplayed)

if searchBar.is_enabled():
    searchBarEnabled = "Search Bar is enabled"
    print(searchBarEnabled)
    buttonEnabled.append(searchBarEnabled)
else:
    searchBarDisabled = "Search Bar is disabled"
    print(searchBarDisabled)
    buttonDisabled = searchBarDisabled

searchBar.click()
searchBarClickAction = "Select search bar passed"
print("Select search bar passed")
clickActionPassed.append(searchBarClickAction)

searchBar.send_keys("Pepsi")

searchBar.send_keys(Keys.RETURN)
searchResultClickAction = "Search result passed"
print(searchResultClickAction)
clickActionPassed.append(searchResultClickAction)

# Hover, favorite, add to cart mute and more information test
thumbnail = driver.find_element(by=By.XPATH, value="/html/body/div[1]/div/div[2]/div[3]/div[3]/div[3]/div[1]")
action = ActionChains(driver)

action.move_to_element(thumbnail).perform()

wait = WebDriverWait(driver, 3)
try:
    mute_button = wait.until(EC.visibility_of_element_located((By.CLASS_NAME, "gvpi-custom-mute-button-mini-hover")))
    allVideosHoverPass = "Hover over all videos thumbnail passed"
    print(allVideosHoverPass)
    hoverActionPassed.append(allVideosHoverPass)
except:
    allVideosHoverFail = "Hover over all videos thumbnail failed"
    print(allVideosHoverFail)
    hoverActionFailed.append((allVideosHoverFail))

# Hover video buttons
mute = driver.find_element(by=By.CLASS_NAME, value="gvpi-custom-mute-button-mini-hover")
favorite = driver.find_element(by=By.CLASS_NAME, value="svg-icon-container")
copy = driver.find_element(by=By.CLASS_NAME, value="fill-icon")
moreInformation = driver.find_element(by=By.CLASS_NAME, value="slider-item-icon-play")

time.sleep(2)

handle_element(mute, "Un-Mute button")
click_and_record(mute, "Un-Mute button", clickActionPassed)
time.sleep(1)
click_and_record(mute, "Mute button", clickActionPassed)

time.sleep(2)

handle_element(favorite, "Favorite button")
click_and_record(favorite, "Favorite button", clickActionPassed)
time.sleep(1)
click_and_record(favorite, "Un-Favorite button", clickActionPassed)
time.sleep(1)

handle_element(copy, "Copy button")

handle_element(moreInformation, "More information button")
click_and_record(moreInformation, "More information button", clickActionPassed)
time.sleep(2)
backMoreInfo = driver.find_element(by=By.CLASS_NAME, value="details-nav-back")
backMoreInfo.click()

time.sleep(2)



# Check for displayed, enabled, clicked, hover actions for navigation bar

# BRANDS
brands = driver.find_element(by=By.XPATH, value="/html/body/div[1]/div/div[2]/div[2]/div[2]/div[2]/p")
handle_element(brands, "Brands Button")
time.sleep(2)
click_and_record(brands, "Brands button", clickActionPassed)

thumbnail = driver.find_element(by=By.XPATH, value="/html/body/div[1]/div/div[2]/div[3]/div[3]/div[3]/div[1]")
action = ActionChains(driver)
action.move_to_element(thumbnail).perform()

wait = WebDriverWait(driver, 3)
try:
    mute_button = wait.until(EC.visibility_of_element_located((By.CLASS_NAME, "gvpi-custom-mute-button-mini-hover")))
    allVideosHoverPass = "Hover over brands videos thumbnail passed"
    print(allVideosHoverPass)
    hoverActionPassed.append(allVideosHoverPass)
except:
    allVideosHoverFail = "Hover over brands videos thumbnail failed"
    print(allVideosHoverFail)
    hoverActionFailed.append(allVideosHoverFail)


mute = driver.find_element(by=By.CLASS_NAME, value="gvpi-custom-mute-button-mini-hover")
favorite = driver.find_element(by=By.CLASS_NAME, value="svg-icon-container")
copy = driver.find_element(by=By.CLASS_NAME, value="fill-icon")
moreInformation = driver.find_element(by=By.CLASS_NAME, value="slider-item-icon-play")

time.sleep(2)

handle_element(mute, "Un-Mute button (brands)")
click_and_record(mute, "Un-Mute button (brands)", clickActionPassed)
time.sleep(1)
click_and_record(mute, "Mute button (brands)", clickActionPassed)

time.sleep(2)

handle_element(favorite, "Favorite button (brands)")
click_and_record(favorite, "Favorite button (brands)", clickActionPassed)
time.sleep(1)
click_and_record(favorite, "Un-Favorite button (brands)", clickActionPassed)
time.sleep(1)

handle_element(copy, "Copy button (brands)")

handle_element(moreInformation, "More information button (brands)")
click_and_record(moreInformation, "More information button (brands)", clickActionPassed)
time.sleep(2)
backMoreInfo = driver.find_element(by=By.CLASS_NAME, value="details-nav-back")
backMoreInfo.click()

time.sleep(3)


#PEOPLE
people = driver.find_element(by=By.XPATH, value="/html/body/div[1]/div/div[2]/div[2]/div[2]/div[3]/p")
handle_element(people, "People button")
time.sleep(2)
click_and_record(people, "People button", clickActionPassed)

thumbnail = driver.find_element(by=By.XPATH, value="/html/body/div[1]/div/div[2]/div[3]/div[3]/div[3]/div[1]")
action = ActionChains(driver)
action.move_to_element(thumbnail).perform()

wait = WebDriverWait(driver, 3)
try:
    mute_button = wait.until(EC.visibility_of_element_located((By.CLASS_NAME, "gvpi-custom-mute-button-mini-hover")))
    allVideosHoverPass = "Hover over people videos thumbnail passed"
    print(allVideosHoverPass)
    hoverActionPassed.append(allVideosHoverPass)
except:
    allVideosHoverFail = "Hover over people videos thumbnail failed"
    print(allVideosHoverFail)
    hoverActionFailed.append(allVideosHoverFail)

mute = driver.find_element(by=By.CLASS_NAME, value="gvpi-custom-mute-button-mini-hover")
favorite = driver.find_element(by=By.CLASS_NAME, value="svg-icon-container")
copy = driver.find_element(by=By.CLASS_NAME, value="fill-icon")
moreInformation = driver.find_element(by=By.CLASS_NAME, value="slider-item-icon-play")

time.sleep(2)

handle_element(mute, "Un-Mute button (people)")
click_and_record(mute, "Un-Mute button (people)", clickActionPassed)
time.sleep(1)
click_and_record(mute, "Mute button (people)", clickActionPassed)

time.sleep(2)

handle_element(favorite, "Favorite button (people)")
click_and_record(favorite, "Favorite button (people)", clickActionPassed)
time.sleep(1)
click_and_record(favorite, "Un-Favorite button (people)", clickActionPassed)
time.sleep(1)

handle_element(copy, "Copy button (people)")

handle_element(moreInformation, "More information button (people)")
click_and_record(moreInformation, "More information button (people)", clickActionPassed)
time.sleep(2)
backMoreInfo = driver.find_element(by=By.CLASS_NAME, value="details-nav-back")
backMoreInfo.click()

time.sleep(3)


transcript = driver.find_element(by=By.XPATH, value="/html/body/div[1]/div/div[2]/div[2]/div[2]/div[4]/p")
handle_element(transcript, "Transcript button")
time.sleep(2)
click_and_record(transcript, "Transcript button", clickActionPassed)

thumbnail = driver.find_element(by=By.XPATH, value="/html/body/div[1]/div/div[2]/div[3]/div[3]/div[3]/div[1]")
action = ActionChains(driver)
action.move_to_element(thumbnail).perform()

wait = WebDriverWait(driver, 3)
try:
    mute_button = wait.until(EC.visibility_of_element_located((By.CLASS_NAME, "gvpi-custom-mute-button-mini-hover")))
    allVideosHoverPass = "Hover over transcript videos thumbnail passed"
    print(allVideosHoverPass)
    hoverActionPassed.append(allVideosHoverPass)
except:
    allVideosHoverFail = "Hover over transcript videos thumbnail failed"
    print(allVideosHoverFail)
    hoverActionFailed.append(allVideosHoverFail)

mute = driver.find_element(by=By.CLASS_NAME, value="gvpi-custom-mute-button-mini-hover")
favorite = driver.find_element(by=By.CLASS_NAME, value="svg-icon-container")
copy = driver.find_element(by=By.CLASS_NAME, value="fill-icon")
moreInformation = driver.find_element(by=By.CLASS_NAME, value="slider-item-icon-play")

time.sleep(2)

handle_element(mute, "Un-Mute button (transcript)")
click_and_record(mute, "Un-Mute button (transcript)", clickActionPassed)
time.sleep(1)
click_and_record(mute, "Mute button (transcript)", clickActionPassed)

time.sleep(2)

handle_element(favorite, "Favorite button (transcript)")
click_and_record(favorite, "Favorite button (transcript)", clickActionPassed)
time.sleep(1)
click_and_record(favorite, "Un-Favorite button (transcript)", clickActionPassed)
time.sleep(1)

handle_element(copy, "Copy button (transcript)")

handle_element(moreInformation, "More information button (transcript))")
click_and_record(moreInformation, "More information button (transcript)", clickActionPassed)
time.sleep(2)
backMoreInfo = driver.find_element(by=By.CLASS_NAME, value="details-nav-back")
backMoreInfo.click()

time.sleep(3)


objects = driver.find_element(by=By.XPATH, value="/html/body/div[1]/div/div[2]/div[2]/div[2]/div[5]/p")
handle_element(objects, "Object button")
time.sleep(2)
click_and_record(objects, "Objects button", clickActionPassed)

thumbnail = driver.find_element(by=By.XPATH, value="/html/body/div[1]/div/div[2]/div[3]/div[3]/div[3]/div[1]")
action = ActionChains(driver)
action.move_to_element(thumbnail).perform()

wait = WebDriverWait(driver, 3)
try:
    mute_button = wait.until(EC.visibility_of_element_located((By.CLASS_NAME, "gvpi-custom-mute-button-mini-hover")))
    allVideosHoverPass = "Hover over object videos thumbnail passed"
    print(allVideosHoverPass)
    hoverActionPassed.append(allVideosHoverPass)
except:
    allVideosHoverFail = "Hover over object videos thumbnail failed"
    print(allVideosHoverFail)
    hoverActionFailed.append(allVideosHoverFail)

mute = driver.find_element(by=By.CLASS_NAME, value="gvpi-custom-mute-button-mini-hover")
favorite = driver.find_element(by=By.CLASS_NAME, value="svg-icon-container")
copy = driver.find_element(by=By.CLASS_NAME, value="fill-icon")
moreInformation = driver.find_element(by=By.CLASS_NAME, value="slider-item-icon-play")

time.sleep(2)

handle_element(mute, "Un-Mute button (objects)")
click_and_record(mute, "Un-Mute button  (objects)", clickActionPassed)
time.sleep(1)
click_and_record(mute, "Mute button  (objects)", clickActionPassed)

time.sleep(2)

handle_element(favorite, "Favorite button  (objects)")
click_and_record(favorite, "Favorite button  (objects)", clickActionPassed)
time.sleep(1)
click_and_record(favorite, "Un-Favorite button (objects)", clickActionPassed)
time.sleep(1)

handle_element(copy, "Copy button (objects)")

handle_element(moreInformation, "More information button (objects)")
click_and_record(moreInformation, "More information button (objects)", clickActionPassed)
time.sleep(2)
backMoreInfo = driver.find_element(by=By.CLASS_NAME, value="details-nav-back")
backMoreInfo.click()

time.sleep(3)
segments = driver.find_element(by=By.CLASS_NAME, value="sm-views-button")
allResults = driver.find_element(by=By.XPATH, value="/html/body/div[1]/div/div[2]/div[3]/div[2]/div/div[1]/div[2]/div[1]")
sortBy = driver.find_element(by=By.CLASS_NAME, value="sort-filter-options")
durationShortest = driver.find_element(by=By.XPATH, value="/html/body/div[1]/div/div[2]/div[3]/div[2]/div/div[2]/div/ul/li[1]")
durationLongest = driver.find_element(by=By.XPATH, value="/html/body/div[1]/div/div[2]/div[3]/div[2]/div/div[2]/div/ul/li[2]")
airDateNewest = driver.find_element(by=By.XPATH, value="/html/body/div[1]/div/div[2]/div[3]/div[2]/div/div[2]/div/ul/li[4]")
airDateOldest = driver.find_element(by=By.XPATH, value="/html/body/div[1]/div/div[2]/div[3]/div[2]/div/div[2]/div/ul/li[5]")
titleAtoZ = driver.find_element(by=By.XPATH, value="/html/body/div[1]/div/div[2]/div[3]/div[2]/div/div[2]/div/ul/li[7]")
titleZtoA = driver.find_element(by=By.XPATH, value="/html/body/div[1]/div/div[2]/div[3]/div[2]/div/div[2]/div/ul/li[8]")
relevance = driver.find_element(by=By.XPATH, value="/html/body/div[1]/div/div[2]/div[3]/div[2]/div/div[2]/div/ul/li[10]")
order = driver.find_element(by=By.CLASS_NAME, value = "icons-sort")




handle_element(allResults, "All Results button")
click_and_record(allResults, "All results button", clickActionPassed)
time.sleep(3)

handle_element(segments, "Segments button")
click_and_record(segments, "Segments button", clickActionPassed)
time.sleep(3)

handle_element(sortBy, "Sort by button")
click_and_record(sortBy, "Sort by button button", clickActionPassed)
time.sleep(3)


sortBy.click()
time.sleep(2)


sortBy.click()
time.sleep(2)

handle_element(durationShortest, "DURATION SHORTEST button")
click_and_record(durationShortest, "DURATION SHORTEST button", clickActionPassed)
time.sleep(3)

sortBy.click()
time.sleep(2)

handle_element(durationLongest, "DURATION LONGEST button")
click_and_record(durationLongest, "DURATION LONGEST button", clickActionPassed)
time.sleep(3)

sortBy.click()
time.sleep(2)

handle_element(airDateNewest, "EVENT DATE NEW button")
click_and_record(airDateNewest, "EVENT DAY NEW button", clickActionPassed)
time.sleep(3)

sortBy.click()
time.sleep(2)

handle_element(airDateOldest, "EVENT DATE OLD button")
click_and_record(airDateOldest, "EVENT DATE OLD button", clickActionPassed)
time.sleep(3)

sortBy.click()
time.sleep(2)

handle_element(titleAtoZ, "TITLE A TO Z button")
click_and_record(titleAtoZ, "TITLE A TO Z button", clickActionPassed)
time.sleep(3)

sortBy.click()
time.sleep(2)

handle_element(titleZtoA, "TITLE Z TO A button")
click_and_record(titleZtoA, "TITLE Z TO A button", clickActionPassed)
time.sleep(3)

sortBy.click()
time.sleep(2)

handle_element(relevance, "RELEVANCE button")
click_and_record(relevance, "RELEVANCE button", clickActionPassed)
time.sleep(3)

order = driver.find_element(by=By.CLASS_NAME, value="icons-sort")
heart = driver.find_element(by=By.CLASS_NAME, value="heart-icon-filter-container")
mic = driver.find_element(by=By.XPATH, value="/html/body/div[1]/div/div[2]/div[1]/div[2]/div[2]")
comboSearch = driver.find_element(by=By.XPATH, value="/html/body/div[1]/div/div[2]/div[1]/div[2]/div[3]")
time.sleep(2)

handle_element(order, "Descending button")
click_and_record(order, "Descending button", clickActionPassed)
time.sleep(3)

handle_element(order, "Ascending button")
click_and_record(order, "Ascending button", clickActionPassed)
time.sleep(3)

handle_element(heart, "Heart button")
click_and_record(heart, "Heart button", clickActionPassed)
time.sleep(3)

handle_element(heart, "Reverse heart button")
click_and_record(heart, "Reverse heart button", clickActionPassed)
time.sleep(3)

handle_element(mic, "Mic button")
click_and_record(mic, "Mic button", clickActionPassed)
time.sleep(3)

handle_element(comboSearch, "Combo Search button")
click_and_record(comboSearch, "Combo Search button", clickActionPassed)
time.sleep(3)

searchBar1DropDown = driver.find_element(by=By.XPATH, value="/html/body/div[1]/div/div[2]/div[1]/div[2]/div[1]/div[1]/div/div[1]")
searchBar1All = driver.find_element(by=By.XPATH, value="/html/body/div[1]/div/div[2]/div[1]/div[2]/div[1]/div[1]/div")
searchBar1Brands = driver.find_element(by=By.XPATH, value="/html/body/div[1]/div/div[2]/div[1]/div[2]/div[1]/div[1]/div/div[2]/div[2]")
searchBar1People = driver.find_element(by=By.XPATH, value="/html/body/div[1]/div/div[2]/div[1]/div[2]/div[1]/div[1]/div/div[2]/div[3]")
searchBar1Transcript = driver.find_element(by=By.XPATH, value="/html/body/div[1]/div/div[2]/div[1]/div[2]/div[1]/div[1]/div/div[2]/div[4]")
searchBar1Objects = driver.find_element(by=By.XPATH, value="/html/body/div[1]/div/div[2]/div[1]/div[2]/div[1]/div[1]/div/div[2]/div[5]")

handle_element(searchBar1DropDown, "Search Bar 1 Drop Down button")
click_and_record(searchBar1DropDown, "Search Bar 1 Drop down button", clickActionPassed)
time.sleep(3)

handle_element(searchBar1All, "Search Bar 1 All button")
click_and_record(searchBar1All, "Search Bar 1 ALl button", clickActionPassed)
time.sleep(3)

searchBar1DropDown.click()
time.sleep(2)

handle_element(searchBar1Brands, "Search Bar 1 Brands button")
click_and_record(searchBar1Brands, "Search Bar 1 Brands button", clickActionPassed)
time.sleep(3)

searchBar1DropDown.click()
time.sleep(2)

handle_element(searchBar1People, "Search Bar 1 People button")
click_and_record(searchBar1People, "Search Bar 1 People button", clickActionPassed)
time.sleep(3)

searchBar1DropDown.click()
time.sleep(2)

handle_element(searchBar1Transcript, "Search Bar 1 Transcript button")
click_and_record(searchBar1Transcript, "Search Bar 1 Transcript button", clickActionPassed)
time.sleep(3)

searchBar1DropDown.click()
time.sleep(2)

handle_element(searchBar1Objects, "Search Bar 1 Objects button")
click_and_record(searchBar1Objects, "Search Bar 1 Objects button", clickActionPassed)
time.sleep(3)

searchBar1 = driver.find_element(by=By.XPATH, value="/html/body/div[1]/div/div[2]/div[1]/div[2]/div[1]/div[1]/input")

handle_element(searchBar1, "Search Bar 1")
click_and_record(searchBar1, "Search Bar 1", clickActionPassed)
time.sleep(3)

andOrDropDown = driver.find_element(by=By.XPATH, value="/html/body/div[1]/div/div[2]/div[1]/div[2]/div[1]/div[2]")
And = driver.find_element(by=By.XPATH, value="/html/body/div[1]/div/div[2]/div[1]/div[2]/div[1]/div[2]/div[2]/div[1]")
Or = driver.find_element(by=By.XPATH, value="/html/body/div[1]/div/div[2]/div[1]/div[2]/div[1]/div[2]/div[2]/div[2]")

handle_element(andOrDropDown, "andOr Drop Down button")
click_and_record(andOrDropDown, "andOr Drop Down button", clickActionPassed)
time.sleep(3)

handle_element(And, "And button")
click_and_record(And, "And button", clickActionPassed)
time.sleep(3)

andOrDropDown.click()
time.sleep(1)

handle_element(Or, "Or button")
click_and_record(Or, "Or button", clickActionPassed)
time.sleep(3)

searchBar2DropDown = driver.find_element(by=By.XPATH, value="/html/body/div[1]/div/div[2]/div[1]/div[2]/div[1]/div[3]/div/div[1]")
searchBar2All = driver.find_element(by=By.XPATH, value="/html/body/div[1]/div/div[2]/div[1]/div[2]/div[1]/div[3]/div/div[2]/div[1]")
searchBar2Brands = driver.find_element(by=By.XPATH, value="/html/body/div[1]/div/div[2]/div[1]/div[2]/div[1]/div[3]/div/div[2]/div[2]")
searchBar2People = driver.find_element(by=By.XPATH, value="/html/body/div[1]/div/div[2]/div[1]/div[2]/div[1]/div[3]/div/div[2]/div[3]")
searchBar2Transcript = driver.find_element(by=By.XPATH, value="/html/body/div[1]/div/div[2]/div[1]/div[2]/div[1]/div[3]/div/div[2]/div[4]")
searchBar2Objects = driver.find_element(by=By.XPATH, value="/html/body/div[1]/div/div[2]/div[1]/div[2]/div[1]/div[3]/div/div[2]/div[5]")

handle_element(searchBar2DropDown, "Search Bar 2 Drop Down button")
click_and_record(searchBar2DropDown, "Search Bar 2 Drop down button", clickActionPassed)
time.sleep(3)

handle_element(searchBar2All, "Search Bar 2 All button")
click_and_record(searchBar2All, "Search Bar 2 ALl button", clickActionPassed)
time.sleep(3)

searchBar2DropDown.click()
time.sleep(2)

handle_element(searchBar2Brands, "Search Bar 2 Brands button")
click_and_record(searchBar2Brands, "Search Bar 2 Brands button", clickActionPassed)
time.sleep(3)

searchBar2DropDown.click()
time.sleep(2)

handle_element(searchBar2People, "Search Bar 2 People button")
click_and_record(searchBar2People, "Search Bar 2 People button", clickActionPassed)
time.sleep(3)

searchBar2DropDown.click()
time.sleep(2)

handle_element(searchBar2Transcript, "Search Bar 2 Transcript button")
click_and_record(searchBar2Transcript, "Search Bar 2 Transcript button", clickActionPassed)
time.sleep(3)

searchBar2DropDown.click()
time.sleep(2)

handle_element(searchBar2Objects, "Search Bar 2 Objects button")
click_and_record(searchBar2Objects, "Search Bar 2 Objects button", clickActionPassed)
time.sleep(3)

searchBar2 = driver.find_element(by=By.XPATH, value="/html/body/div[1]/div/div[2]/div[1]/div[2]/div[1]/div[3]/input")

handle_element(searchBar2, "Search Bar 2")
click_and_record(searchBar2, "Search Bar 2", clickActionPassed)
time.sleep(3)

filter = driver.find_element(by=By.XPATH, value="/html/body/div[1]/div/div[2]/div[3]/div[1]/div/div")
category = driver.find_element(by=By.XPATH, value="/html/body/div[1]/div/div[2]/div[3]/div[1]/div/nav/div[2]/div/div[1]/div")
sports = driver.find_element(by=By.XPATH, value="/html/body/div[1]/div/div[2]/div[3]/div[1]/div/nav/div[2]/div/div[1]/label[1]")
news = driver.find_element(by=By.XPATH, value="/html/body/div[1]/div/div[2]/div[3]/div[1]/div/nav/div[2]/div/div[1]/label[2]")
weather = driver.find_element(by=By.XPATH, value="/html/body/div[1]/div/div[2]/div[3]/div[1]/div/nav/div[2]/div/div[1]/label[3]")
nature = driver.find_element(by=By.XPATH, value="/html/body/div[1]/div/div[2]/div[3]/div[1]/div/nav/div[2]/div/div[1]/label[4]")
TVShow = driver.find_element(by=By.XPATH, value="/html/body/div[1]/div/div[2]/div[3]/div[1]/div/nav/div[2]/div/div[1]/label[5]")
documentary = driver.find_element(by=By.XPATH, value="/html/body/div[1]/div/div[2]/div[3]/div[1]/div/nav/div[2]/div/div[1]/label[6]")

handle_element(filter, "Filter Button")
click_and_record(filter, "Filter Button", clickActionPassed)
time.sleep(2)

handle_element(category, "Category Button")
click_and_record(category, "Category Button", clickActionPassed)
time.sleep(1)

handle_element(sports, "Sports Button")
click_and_record(sports, "Sports Button", clickActionPassed)
time.sleep(1)

handle_element(news, "News Button")
click_and_record(news, "News Button", clickActionPassed)
time.sleep(1)

handle_element(weather, "Weather Button")
click_and_record(weather, "Weather Button", clickActionPassed)
time.sleep(1)

handle_element(nature, "Nature Button")
click_and_record(nature, "Nature Button", clickActionPassed)
time.sleep(1)

handle_element(TVShow, "TV Show Button")
click_and_record(TVShow, "TV Show Button", clickActionPassed)
time.sleep(1)

handle_element(documentary, "Documentary Button")
click_and_record(documentary, "Documentary Button", clickActionPassed)
time.sleep(1)


handle_element(sports, "Sports Button deselect")
click_and_record(sports, "Sports Button deselect", clickActionPassed)
time.sleep(1)

handle_element(news, "News Button deselect")
click_and_record(news, "News Button deselect", clickActionPassed)
time.sleep(1)

handle_element(weather, "Weather Button deselect")
click_and_record(weather, "Weather Button deselect", clickActionPassed)
time.sleep(1)

handle_element(nature, "Nature Button deselect")
click_and_record(nature, "Nature Button deselect", clickActionPassed)
time.sleep(1)

handle_element(TVShow, "TV Show Button deselect")
click_and_record(TVShow, "TV Show Button deselect", clickActionPassed)
time.sleep(1)

handle_element(documentary, "Documentary Button deselect")
click_and_record(documentary, "Documentary Button deselect", clickActionPassed)
time.sleep(1)

handle_element(category, "Category Button deselect")
click_and_record(category, "Category Button deselect", clickActionPassed)
time.sleep(1)

years = driver.find_element(by=By.XPATH, value="/html/body/div[1]/div/div[2]/div[3]/div[1]/div/nav/div[2]/div/div[2]/div")

handle_element(years, "Years Button")
click_and_record(years, "Years Button", clickActionPassed)
time.sleep(1)

year_elements = {
    "yeara": ("2020-2023 Button", "/html/body/div[1]/div/div[2]/div[3]/div[1]/div/nav/div[2]/div/div[2]/label[1]"),
    "yearb": ("2010-2019 Button", "/html/body/div[1]/div/div[2]/div[3]/div[1]/div/nav/div[2]/div/div[2]/label[2]"),
    "yearc": ("2000-2009 Button", "/html/body/div[1]/div/div[2]/div[3]/div[1]/div/nav/div[2]/div/div[2]/label[3]"),
    "yeard": ("1990-1999 Button", "/html/body/div[1]/div/div[2]/div[3]/div[1]/div/nav/div[2]/div/div[2]/label[4]"),
    "yeare": ("1980-1989 Button", "/html/body/div[1]/div/div[2]/div[3]/div[1]/div/nav/div[2]/div/div[2]/label[5]"),
    "yearf": ("1970-1979 Button", "/html/body/div[1]/div/div[2]/div[3]/div[1]/div/nav/div[2]/div/div[2]/label[6]"),
    "yearg": ("1960-1969 Button", "/html/body/div[1]/div/div[2]/div[3]/div[1]/div/nav/div[2]/div/div[2]/label[7]"),
    "yearh": ("1950-1959 Button", "/html/body/div[1]/div/div[2]/div[3]/div[1]/div/nav/div[2]/div/div[2]/label[8]"),
}

# Loop through the dictionary and handle each element
for name, (label, xpath) in year_elements.items():
    elem = driver.find_element(by=By.XPATH, value=xpath)
    handle_element(elem, label)
    click_and_record(elem, label, clickActionPassed)
    time.sleep(1)

year_elements_deselect = {
    "yeara": ("2020-2023 Button Deselect", "/html/body/div[1]/div/div[2]/div[3]/div[1]/div/nav/div[2]/div/div[2]/label[1]"),
    "yearb": ("2010-2019 Button Deselect", "/html/body/div[1]/div/div[2]/div[3]/div[1]/div/nav/div[2]/div/div[2]/label[2]"),
    "yearc": ("2000-2009 Button Deselect", "/html/body/div[1]/div/div[2]/div[3]/div[1]/div/nav/div[2]/div/div[2]/label[3]"),
    "yeard": ("1990-1999 Button Deselect", "/html/body/div[1]/div/div[2]/div[3]/div[1]/div/nav/div[2]/div/div[2]/label[4]"),
    "yeare": ("1980-1989 Button Deselect", "/html/body/div[1]/div/div[2]/div[3]/div[1]/div/nav/div[2]/div/div[2]/label[5]"),
    "yearf": ("1970-1979 Button Deselect", "/html/body/div[1]/div/div[2]/div[3]/div[1]/div/nav/div[2]/div/div[2]/label[6]"),
    "yearg": ("1960-1969 Button Deselect", "/html/body/div[1]/div/div[2]/div[3]/div[1]/div/nav/div[2]/div/div[2]/label[7]"),
    "yearh": ("1950-1959 Button Deselect", "/html/body/div[1]/div/div[2]/div[3]/div[1]/div/nav/div[2]/div/div[2]/label[8]"),
}

# Loop through the dictionary and handle each element
for name, (label, xpath) in year_elements_deselect.items():
    elem = driver.find_element(by=By.XPATH, value=xpath)
    handle_element(elem, label)
    click_and_record(elem, label, clickActionPassed)
    time.sleep(1)

handle_element(years, "Years Button Deselect")
click_and_record(years, "Years Button Deselect", clickActionPassed)
time.sleep(1)

filter = driver.find_element(by=By.XPATH, value="/html/body/div[1]/div/div[2]/div[3]/div[1]/div/div")
date = driver.find_element(by=By.XPATH, value="/html/body/div[1]/div/div[2]/div[3]/div[1]/div/nav/div[2]/div/div[3]/div[1]")
startDate = driver.find_element(by=By.XPATH, value="/html/body/div[1]/div/div[2]/div[3]/div[1]/div/nav/div[2]/div/div[3]/div[2]/div[1]")
startCalendar = driver.find_element(by=By.XPATH, value="/html/body/div[1]/div/div[2]/div[3]/div[1]/div/nav/div[2]/div/div[3]/div[2]/div[3]")
endDate = driver.find_element(by=By.XPATH, value="/html/body/div[1]/div/div[2]/div[3]/div[1]/div/nav/div[2]/div/div[3]/div[3]/div[1]")
endCalendar = driver.find_element(by=By.XPATH, value="/html/body/div[1]/div/div[2]/div[3]/div[1]/div/nav/div[2]/div/div[3]/div[3]/div[3]")


handle_element(date, "Date Button")
click_and_record(date, "Date Button", clickActionPassed)
time.sleep(2)

# Initialize WebDriverWait with a timeout of 3 seconds
wait = WebDriverWait(driver, 3)

# Click on the Start Date Button
handle_element(startDate, "Start Date Button")
click_and_record(startDate, "Start Date Button", clickActionPassed)
time.sleep(2)

# Wait for the startCalendar to become visible
try:
    startCalendar = wait.until(EC.visibility_of_element_located((By.XPATH, "/html/body/div[1]/div/div[2]/div[3]/div[1]/div/nav/div[2]/div/div[3]/div[2]/div[3]")))
    startCalendarVisibleSuccess = "Start date button passed."
    print(startCalendarVisibleSuccess)
    clickActionPassed.append(startCalendarVisibleSuccess)
except:
    startCalendarVisibleFail = "Start date button failed."
    print(startCalendarVisibleFail)
    clickActionFailed.append(startCalendarVisibleFail)

handle_element(startDate, "Start Date Button Deselect")
click_and_record(startDate, "Start Date Button Deselect", clickActionPassed)
time.sleep(2)

# Initialize WebDriverWait with a timeout of 3 seconds
wait = WebDriverWait(driver, 3)

# Click on the End Date Button
handle_element(endDate, "End Date Button")
click_and_record(endDate, "End Date Button", clickActionPassed)
time.sleep(2)

# Wait for the endCalendar to become visible
try:
    endCalendar = wait.until(EC.visibility_of_element_located((By.XPATH, "/html/body/div[1]/div/div[2]/div[3]/div[1]/div/nav/div[2]/div/div[3]/div[3]/div[3]")))
    endCalendarVisibleSuccess = "End date button passed"
    print(endCalendarVisibleSuccess)
    clickActionPassed.append(endCalendarVisibleSuccess)
except:
    endCalendarVisibleFail = "End date button did not pass"
    print(endCalendarVisibleFail)
    clickActionFailed.append(endCalendarVisibleFail)

handle_element(endDate, "End Date Button Deselect")
click_and_record(endDate, "End Date Button Deselect", clickActionPassed)
time.sleep(2)

duration = driver.find_element(by=By.XPATH, value="/html/body/div[1]/div/div[2]/div[3]/div[1]/div/nav/div[2]/div/div[4]/div[1]")
slider = driver.find_element(by=By.XPATH, value="/html/body/div[1]/div/div[2]/div[3]/div[1]/div/nav/div[2]/div/div[4]/div[2]/input")

handle_element(duration, "Duration Button")
click_and_record(duration, "Duration Button", clickActionPassed)
time.sleep(2)

ActionChains(driver).drag_and_drop_by_offset(slider, -60, 0).perform()
time.sleep(2)
sliderMovement = "Slider movement passed"
print(sliderMovement)
clickActionPassed.append(sliderMovement)
time.sleep(2)

handle_element(duration, "Duration Button Deselect")
click_and_record(duration, "Duration Button Deselect", clickActionPassed)
time.sleep(2)

filter.click()

ws['A1'] = "Click Actions Passed"
ws['B1'] = "Buttons Displayed"
ws['C1'] = "Buttons Enabled"
ws['D1'] = "Hover action passed"
ws['E1'] = "Buttons Not Displayed"
ws['F1'] = "Buttons Disabled"
ws['G1'] = "Hover Action Failed"
ws['H1'] = "Click Action Failed"

for index, value in enumerate(clickActionPassed):
    cell = ws.cell(row=index+2, column=1, value=value)

for index, value in enumerate(buttonDisplayed):
    cell = ws.cell(row=index+2, column=2, value=value)

for index, value in enumerate(buttonEnabled):
    cell = ws.cell(row=index+2, column=3, value=value)

for index, value in enumerate(hoverActionPassed):
    cell = ws.cell(row=index+2, column=4, value=value)

for index, value in enumerate(buttonNotDisplayed):
    cell = ws.cell(row=index+2, column=5, value=value)

for index, value in enumerate(buttonDisabled):
    cell = ws.cell(row=index+2, column=6, value=value)

for index, value in enumerate(hoverActionFailed):
    cell = ws.cell(row=index+2, column=7, value=value)

for index, value in enumerate(hoverActionFailed):
    cell = ws.cell(row=index+2, column=8, value=value)

wb.save('SearchPage.xlsx')