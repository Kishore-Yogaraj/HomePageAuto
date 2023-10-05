from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.action_chains import ActionChains
from openpyxl import Workbook
import pyautogui
import time
import subprocess
import os
import re
from strip_ansi import strip_ansi

def returnToHomeandCat():
    time.sleep(2)
    logo = driver.find_element(by=By.CSS_SELECTOR, value='[alt="banner"]')
    logo.click()
    time.sleep(2)
    categories = driver.find_element(by=By.CLASS_NAME, value="nav-item-child")
    categories.click()
    time.sleep(3)

def returnToHome():
    time.sleep(2)
    logo = driver.find_element(by=By.CSS_SELECTOR, value='[alt="banner"]')
    if logo.is_displayed():
       print("Logo return to home is displayed")
    else:
        print("Logo return to home is not displayed")
    logo.click()
    time.sleep(2)


def randomClick():
    pyautogui.moveTo(200, 200)
    pyautogui.click()

def profileIcon():
    profileIcon = driver.find_element(by=By.XPATH, value="/html/body/div[1]/div/div[1]/div[2]/div[3]")
    if profileIcon.is_displayed():
        print("Profile Icon button is displayed")
    else:
        print("Profile Icon button is not displayed")
    profileIcon.click()
    print("Profile Icon passed")
    time.sleep(1)

def run_newman_and_save_output(newman_command, folder_name, output_file, error_file):
    # Add Node's directory to PATH
    os.environ["PATH"] += os.pathsep + r"C:\Eon\NodeJS"

    # Run Newman command and capture the output
    result = subprocess.run(
        newman_command,
        shell=True,
        stdout=subprocess.PIPE,
        stderr=subprocess.PIPE,
        encoding="utf-8",
        text=True,
    )

    # Define a regular expression pattern to match the relevant output
    pattern = f"{folder_name}.*?(?=┌|$)"  # This pattern captures the output for the specified folder

    # Search for the pattern in the Newman output
    match = re.search(pattern, result.stdout, re.DOTALL)

    if match:
        # Extract and clean the matched output
        extracted_output = match.group(0).strip()

        # Use strip_ansi to remove ANSI escape codes
        cleaned_output = strip_ansi(extracted_output)

        # Split the cleaned output into lines
        lines = cleaned_output.split("\n")

        # Create a list to store lines with checkmarks and lines with numbers
        checkmark_lines = []
        number_lines = []

        for line in lines:
            if re.match(r'^\s*√', line):
                checkmark_lines.append(line)
            elif re.match(r'^\s*\d+\.', line):
                number_lines.append(line)
            else:
                checkmark_lines.append(line)

        # Append the cleaned output to the specified output file
        with open(output_file, "a", encoding="utf-8") as file:
            file.write("\n".join(checkmark_lines))

            # Append the number lines to the error file with appropriate headings
            with open(error_file, "a", encoding="utf-8") as file:
                current_heading = None
                wrote_folder_name = False  # Flag to check if folder_name has been written to the file

                for line in number_lines:
                    match = re.match(r'^\s*(\d+\.)\s*(.*)', line)
                    if match:
                        # Extract the heading and the content
                        heading, content = match.groups()
                        # If the folder_name hasn't been written yet, write it
                        if not wrote_folder_name:
                            # Writing folder_name followed by a newline character
                            file.write(f"{folder_name}\n")
                            wrote_folder_name = True
                        file.write(f"{heading} {content}\n")
                    else:
                        # If there's no valid heading, just write the line as is
                        file.write(line + "\n")
                # Insert an additional newline after writing all errors for the current folder
                file.write("\n")
            # ... [rest of your function after this section]


    else:
        print(f"No relevant output found for {folder_name}")

    print()

    # Print the cleaned output
    print("\n".join(checkmark_lines))

output_file = "combined_output.txt"
error_file = "error_output.txt"

def run_newman_for_folder(folder_name, output_filename='error_output.txt'):
    # Ensure Node's directory is in PATH
    os.environ["PATH"] += os.pathsep + r"C:\Eon\NodeJS"

    newman_path = r"C:\Users\kisho\AppData\Roaming\npm\newman"
    newman_command = f"{newman_path} run \"C:\Eon\JSON File\LandingPage.postman_collection.json\" --folder {folder_name}"
    result = subprocess.run(newman_command, shell=True, stdout=subprocess.PIPE, stderr=subprocess.PIPE, encoding='utf-8', text=True)

    # Define a regular expression pattern to match the assertion failure details section
    pattern = r"\d+\.\s+AssertionError.*?(?=\d+\.\s+AssertionError|\Z)"

    # Search for the pattern in the Newman output
    matches = re.findall(pattern, result.stdout, re.DOTALL)

    # Strip ANSI escape sequences from the matches
    ansi_escape = re.compile(r'\x1B\[[0-?]*[ -/]*[@-~]')
    stripped_matches = [ansi_escape.sub('', match) for match in matches]

    # Prepare the output
    if stripped_matches:
        failure_details_output = '\n'.join(stripped_matches)

        # Write the output to a text file
        with open(output_filename, 'a', encoding='utf-8') as file:  # 'a' mode for appending
            file.write(failure_details_output + '\n\n')  # Added extra newline for separation

        # Print to console
        print(failure_details_output)


newman_path = r"C:\Users\kisho\AppData\Roaming\npm\newman"


wb = Workbook()

ws = wb.active

clickActionPassed = []
hoverActionPassed = []
buttonDisplayed = []
buttonEnabled = []
buttonNotDisplayed = []
buttonDisabled =[]



# Create an Options object to store Chrome options
options = Options()

# Window size
options.add_argument("--window-size=1920,1080")

# This allows the Chrome window to remain open after the script finishes
options.add_experimental_option("detach", True)

# options.add_argument("--headless=new")

# Initialize the Chrome driverai-user   vq4alm56

driver = webdriver.Chrome(options=options)

# Access link to website
url = "https://wralstaging.eonmedia.ai/"
driver.get(url)

# Wait for pop up to open
time.sleep(2)

# Type the username and press TAB to switch to the password field
pyautogui.typewrite('ai-user')
pyautogui.press('tab')

# Type the password and press Enter to submit
pyautogui.typewrite('vq4alm56')

pyautogui.press('enter')

time.sleep(3)

login = driver.find_element(by=By.ID, value="log-in-button")

newman_command_sso_google_login = f"{newman_path} run \"C:\Eon\JSON File\LandingPage.postman_collection.json\" --folder sso_google_login"
run_newman_and_save_output(newman_command_sso_google_login, "sso_google_login", output_file, error_file)
run_newman_for_folder('sso_google_login')

newman_command_login = f"{newman_path} run \"C:\Eon\JSON File\LandingPage.postman_collection.json\" --folder login"
run_newman_and_save_output(newman_command_login, "login", output_file, error_file)
run_newman_for_folder('login')

if login.is_displayed():
    logInButtonDisplayed = "Login is displayed"
    print(logInButtonDisplayed)
    buttonDisplayed.append(logInButtonDisplayed)
else:
    logInButtonNotDisplayed = "Login is not displayed"
    print(logInButtonNotDisplayed)
    buttonNotDisplayed.append(logInButtonNotDisplayed)

if login.is_enabled():
    logInButtonEnabled = "Log In button is enabled"
    print(logInButtonEnabled)
    buttonEnabled.append(logInButtonEnabled)
else:
    logInButtonDisabled = "Log In button is disabled"
    print(logInButtonDisabled)
    buttonDisabled.append(logInButtonDisabled)

login.click()
logInPass = "Log In Passed"
print(logInPass)
clickActionPassed.append(logInPass)


time.sleep(3)

emailInput = driver.find_element(by=By.XPATH, value="/html/body/div[1]/div/div[1]/div[6]/div/div[1]/form/input")
if emailInput.is_displayed():
    emailInputDisplayed = "Email input text box is displayed"
    print(emailInputDisplayed)
    buttonDisplayed.append((emailInputDisplayed))
else:
    emailInputNotDisplayed = "Email input text box is not displayed"
    print(emailInputNotDisplayed)
    buttonNotDisplayed.append(emailInputNotDisplayed)

emailInput.click()
emailInput.send_keys("kishore.yogaraj@gmail.com")
emailInput.send_keys(Keys.TAB)
emailInputClickAction = "Username passed"
print(emailInputClickAction)
clickActionPassed.append(emailInputClickAction)

time.sleep(1)

passwordInput = driver.find_element(by=By.XPATH, value="/html/body/div[1]/div/div[1]/div[6]/div/div[1]/form/div[3]/input")
if passwordInput.is_displayed():
    passwordInputDisplayed = "Password input text box is displayed"
    print(passwordInputDisplayed)
    buttonDisplayed.append(passwordInputDisplayed)

else:
    passwordInputNotDisplayed = "Password input text box is not displayed"
    print(passwordInputNotDisplayed)
    buttonNotDisplayed.append(passwordInputNotDisplayed)


passwordInput.click()
passwordInput.send_keys("password1")
passwordInputClickAction = "Password passed"
print(passwordInputClickAction)
clickActionPassed.append(passwordInputClickAction)

time.sleep(1)

hidePassword = driver.find_element(by=By.XPATH, value="/html/body/div[1]/div/div[1]/div[6]/div/div[1]/form/div[3]/div")
if hidePassword.is_displayed():
    hidePasswordDisplayed = "Hide password button is displayed"
    print(hidePasswordDisplayed)
    buttonDisplayed.append((hidePasswordDisplayed))
else:
    hidePasswordNotDisplayed = "Hide button password is not displayed"
    print(hidePasswordNotDisplayed)
    buttonNotDisplayed.append(hidePasswordNotDisplayed)

if hidePassword.is_enabled():
    hidePasswordEnabled = "Hide password button is enabled"
    print(hidePasswordEnabled)
    buttonEnabled.append(hidePasswordEnabled)
else:
    hidePasswordDisabled = "Hide password button is disabled"
    print(hidePasswordDisabled)
    buttonEnabled.append(hidePasswordDisabled)

hidePassword.click()

viewPasswordClickAction = "View password successful"
print(viewPasswordClickAction)
clickActionPassed.append(viewPasswordClickAction)
time.sleep(2)

hidePassword.click()

hidePasswordClickAction = "Hide password successful"
print(hidePasswordClickAction)
clickActionPassed.append(hidePasswordClickAction)

time.sleep(1)

signIn = driver.find_element(by=By.XPATH, value="/html/body/div[1]/div/div[1]/div[6]/div/div[1]/form/div[4]/button")
if signIn.is_displayed():
    signInDisplayed = "Sign in button is displayed"
    print(signInDisplayed)
    buttonDisplayed.append((signInDisplayed))
else:
    signInNotDisplayed = "Sign in button is not displayed"
    print(signInNotDisplayed)
    buttonNotDisplayed.append(signInNotDisplayed)

if signIn.is_enabled():
    signInEnabled = "Sign In button is enabled"
    print(signInEnabled)
    buttonEnabled.append(signInEnabled)
else:
    signInDisabled = "Sign In button is disabled"
    print(signInDisabled)
    buttonDisabled.append(signInDisabled)

signIn.click()
signInClickAction = "Sign in button passed"
print(signInClickAction)
clickActionPassed.append(signInClickAction)

loginProcessClickAction = "Login process passed"
print(loginProcessClickAction)
clickActionPassed.append((loginProcessClickAction))

time.sleep(3)

# Profile icon drop down
profile = driver.find_element(by=By.XPATH, value="/html/body/div[1]/div/div[1]/div[2]/div[3]/div[3]/div[1]/div/a[1]")
myOrders = driver.find_element(by=By.XPATH, value="/html/body/div[1]/div/div[1]/div[2]/div[3]/div[3]/div[2]/a[1]/div[2]")
aiDashboard = driver.find_element(by=By.XPATH, value="/html/body/div[1]/div/div[1]/div[2]/div[3]/div[3]/div[2]/a[2]")
manageStorefront = driver.find_element(by=By.XPATH, value="/html/body/div[1]/div/div[1]/div[2]/div[3]/div[3]/div[2]/a[3]")
settings = driver.find_element(by=By.XPATH, value="/html/body/div[1]/div/div[1]/div[2]/div[3]/div[3]/div[2]/a[4]")
logout = driver.find_element(by=By.XPATH, value="/html/body/div[1]/div/div[1]/div[2]/div[3]/div[3]/div[2]/div[3]")

profileIcon()

if profile.is_displayed():
    profileDisplayed = "Profile Icon is displayed"
    print(profileDisplayed)
    buttonDisplayed.append(profileDisplayed)
else:
    profileNotDisplayed = "Profile Icon is not displayed"
    print(profileNotDisplayed)
    buttonNotDisplayed.append(profileNotDisplayed)

if profile.is_enabled():
    profileEnabled = "Profile button is enabled"
    print(profileEnabled)
    buttonEnabled.append(profileEnabled)
else:
    profileDisabled = "Profile button is disabled"
    print(profileDisabled)
    buttonDisabled.append(profileDisabled)

if myOrders.is_displayed():
    myOrdersDisplayed = "My Orders button is displayed"
    print(myOrdersDisplayed)
    buttonDisplayed.append(myOrdersDisplayed)
else:
    myOrdersNotDisplayed = "My Orders button is not displayed"
    print(myOrdersNotDisplayed)
    buttonNotDisplayed.append(myOrdersNotDisplayed)

if myOrders.is_enabled():
    myOrdersEnabled = "My Orders button is enabled"
    print(myOrdersEnabled)
    buttonEnabled.append(myOrdersEnabled)
else:
    myOrdersDisabled = "My Orders button is disabled"
    print(myOrdersDisabled)
    buttonDisabled.append(myOrdersDisabled)

if aiDashboard.is_displayed():
    aiDashboardDisplayed = "AI Dashboard button is displayed"
    print(aiDashboardDisplayed)
    buttonDisplayed.append(aiDashboardDisplayed)
else:
    aiDashboardNotDisplayed = "AI Dashboard button is not displayed"
    print(aiDashboardNotDisplayed)
    buttonNotDisplayed.append(aiDashboardNotDisplayed)

if aiDashboard.is_enabled():
    aiDashboardEnabled = "AI Dashboard button is enabled"
    print(aiDashboardEnabled)
    buttonEnabled.append(aiDashboardEnabled)
else:
    aiDashboardDisabled = "AI Dashboard button is disabled"
    print(aiDashboardDisabled)
    buttonDisabled.append(aiDashboardDisabled)

if manageStorefront.is_displayed():
    manageStorefrontDisplayed = "Manage Storefront button is displayed"
    print(manageStorefrontDisplayed)
    buttonDisplayed.append(manageStorefrontDisplayed)
else:
    manageStorefrontNotDisplayed = "Manage Storefront button is not displayed"
    print(manageStorefrontNotDisplayed)
    buttonNotDisplayed.append(manageStorefrontNotDisplayed)

if manageStorefront.is_enabled():
    manageStorefrontEnabled = "Manage Storefront button is enabled"
    print(manageStorefrontEnabled)
    buttonEnabled.append(manageStorefrontEnabled)
else:
    manageStorefrontDisabled = "Manage Storefront button is disabled"
    print(manageStorefrontDisabled)
    buttonDisabled.append(manageStorefrontDisabled)

newman_command_setting = f"{newman_path} run \"C:\Eon\JSON File\LandingPage.postman_collection.json\" --folder setting"
run_newman_and_save_output(newman_command_setting, "setting", output_file, error_file)
run_newman_for_folder('setting')

if settings.is_displayed():
    settingDisplayed = "Settings button is displayed"
    print(settingDisplayed)
    buttonDisplayed.append(settingDisplayed)
else:
    settingsNotDisplayed = "Settings button is not displayed"
    print(settingsNotDisplayed)
    buttonNotDisplayed.append(settingsNotDisplayed)

if settings.is_enabled():
    settingsEnabled = "Settings button is enabled"
    print(settingsEnabled)
    buttonEnabled.append(settingsEnabled)
else:
    settingDisabled = "Setting button is disabled"
    print(settingDisabled)
    buttonDisabled.append(settingDisabled)

if logout.is_displayed():
    logoutDisplayed = "Logout button is displayed"
    print(logoutDisplayed)
    buttonDisplayed.append(logoutDisplayed)
else:
    logoutNotDisplayed = "Logout button is not displayed"
    print(logoutNotDisplayed)
    buttonNotDisplayed.append(logoutNotDisplayed)

if logout.is_enabled():
    logoutEnabled = "Logout button is enabled"
    print(logoutEnabled)
    buttonEnabled.append(logoutEnabled)
else:
    logoutDisabled = "Logout button is disabled"
    print(logoutDisabled)
    buttonDisabled.append(logoutDisabled)

randomClick()

newman_command_adding_cart= f"{newman_path} run \"C:\Eon\JSON File\LandingPage.postman_collection.json\" --folder adding_cart"
run_newman_and_save_output(newman_command_adding_cart, "adding_cart", output_file, error_file)
run_newman_for_folder('adding_cart')

newman_command_orders_and_carts= f"{newman_path} run \"C:\Eon\JSON File\LandingPage.postman_collection.json\" --folder orders_and_carts"
run_newman_and_save_output(newman_command_orders_and_carts, "orders_and_carts", output_file, error_file)
run_newman_for_folder('orders_and_carts')

shoppingCartNav = driver.find_element(by=By.XPATH, value="/html/body/div[1]/div/div[1]/div[2]/div[2]")
if shoppingCartNav.is_displayed():
    shoppingCartNavDisplayed = "Shopping cart button is displayed"
    print(shoppingCartNavDisplayed)
    buttonDisplayed.append(shoppingCartNavDisplayed)
else:
    shoppingCartNavNotDisplayed = "Shopping cart button is not displayed"
    print(shoppingCartNavNotDisplayed)
    buttonNotDisplayed.append(shoppingCartNavNotDisplayed)

if shoppingCartNav.is_enabled():
    shoppingCartNavEnabled = "Shopping cart button is enabled"
    print(shoppingCartNavEnabled)
    buttonEnabled.append(shoppingCartNavEnabled)
else:
    shoppingCartNavDisabled = "Shopping cart button is disabled"
    print(shoppingCartNavDisabled)
    buttonDisabled.append(shoppingCartNavDisabled)

shoppingCartNav.click()
shoppingCartNavClickAction = "Shopping cart icon passed"
print(shoppingCartNavClickAction)
clickActionPassed.append(shoppingCartNavClickAction)



returnToHome()

profileIcon()

profile.click()
profileClickAction = "Profile passed"
print(profileClickAction)
clickActionPassed.append(profileClickAction)

returnToHome()

profileIcon()

myOrders.click()
myOrderClickAction = "My orders passed"
print(myOrderClickAction)
clickActionPassed.append(myOrderClickAction)

returnToHome()

profileIcon()

newman_command_ai_dashboard= f"{newman_path} run \"C:\Eon\JSON File\LandingPage.postman_collection.json\" --folder AI_Dashboard"
run_newman_and_save_output(newman_command_ai_dashboard, "AI_Dashboard", output_file, error_file)
run_newman_for_folder('AI_Dashboard')

aiDashboard.click()
aiDashboardClickAction = "AI Dashboard passed"
print(aiDashboardClickAction)
clickActionPassed.append(aiDashboardClickAction)


returnToHome()

profileIcon()

manageStorefront.click()
manageStorefrontClickAction = "Manage Storefront passed"
print(manageStorefrontClickAction)
clickActionPassed.append(manageStorefrontClickAction)

returnToHome()

profileIcon()

settings.click()
settingClickAction = "Settings passed"
print(settingClickAction)
clickActionPassed.append(settingClickAction)

returnToHome()

about = driver.find_element(by=By.LINK_TEXT, value="About")

newman_command_AboutUs = f"{newman_path} run \"C:\Eon\JSON File\LandingPage.postman_collection.json\" --folder AboutUs"
run_newman_and_save_output(newman_command_AboutUs, "AboutUs", output_file, error_file)
run_newman_for_folder('AboutUs')

if about.is_displayed():
    aboutDisplayed = "About button is displayed"
    print(aboutDisplayed)
    buttonDisplayed.append(aboutDisplayed)
else:
    aboutNotDisplayed = "About button is not displayed"
    print(aboutNotDisplayed)
    buttonNotDisplayed.append(aboutNotDisplayed)

if about.is_enabled():
    aboutEnabled = "About button is enabled"
    print(aboutEnabled)
    buttonEnabled.append(aboutEnabled)
else:
    aboutDisabled = "About button is disabled"
    print(aboutDisabled)
    buttonDisabled.append(aboutDisabled)

about.click()
aboutClickAction = "About button passed"
print(aboutClickAction)
clickActionPassed.append(aboutClickAction)



time.sleep(3)

returnToHome()

time.sleep(3)

faq = driver.find_element(by=By.LINK_TEXT, value="FAQ")

newman_command_faq= f"{newman_path} run \"C:\Eon\JSON File\LandingPage.postman_collection.json\" --folder FAQ"
run_newman_and_save_output(newman_command_faq, "FAQ", output_file, error_file)
run_newman_for_folder('FAQ')

if faq.is_displayed():
    faqDisplayed = "FAQ button is displayed"
    print(faqDisplayed)
    buttonDisplayed.append(faqDisplayed)
else:
    faqNotDisplayed = "FAQ button is not displayed"
    print(faqNotDisplayed)
    buttonNotDisplayed.append(faqNotDisplayed)

if faq.is_enabled():
    faqEnabled = "FAQ button is enabled"
    print(faqEnabled)
    buttonEnabled.append(faqEnabled)
else:
    faqDisabled = "FAQ button is disabled"
    print(faqDisabled)
    buttonDisabled.append(faqDisabled)

faq.click()
faqClickAction = "FAQ button passed"
print(faqClickAction)
clickActionPassed.append(faqClickAction)



time.sleep(3)

returnToHome()

time.sleep(3)

categories = driver.find_element(by=By.CLASS_NAME, value="nav-item-child")
if categories.is_displayed():
    categoriesDisplayed = "Categories button is displayed"
    print(categoriesDisplayed)
    buttonDisplayed.append(categoriesDisplayed)
else:
    categoriesNotDisplayed = "Categories button is not displayed"
    print(categoriesNotDisplayed)
    buttonNotDisplayed.append(categoriesNotDisplayed)

if categories.is_enabled():
    categoriesEnabled = "Categories button is enabled"
    print(categoriesEnabled)
    buttonEnabled.append(categoriesEnabled)
else:
    categoriesDisabled = "Categories button is disabled"
    print(categoriesDisabled)
    buttonDisabled.append(categoriesDisabled)

categories.click()
categoriesClickAction = "Categories dropdown passed"
print(categoriesClickAction)
clickActionPassed.append(categoriesClickAction)

time.sleep(3)

# Categories drop down
allShows = driver.find_element(by=By.LINK_TEXT, value="ALL SHOWS")
documentary = driver.find_element(by=By.LINK_TEXT, value="DOCUMENTARY")
news = driver.find_element(by=By.LINK_TEXT, value="NEWS")
sports = driver.find_element(by=By.LINK_TEXT, value="SPORTS")
weather = driver.find_element(by=By.LINK_TEXT, value="WEATHER")
nature = driver.find_element(by=By.LINK_TEXT, value="NATURE")
tvShows = driver.find_element(by=By.LINK_TEXT, value="TV SHOWS")

if allShows.is_displayed():
    allShowsDisplayed = "All Shows button is displayed"
    print(allShowsDisplayed)
    buttonDisplayed.append(allShowsDisplayed)
else:
    allShowsNotDisplayed = "All shows button is not displayed"
    print(allShowsNotDisplayed)
    buttonNotDisplayed.append(allShowsNotDisplayed)

if allShows.is_enabled():
    allShowsEnabled = "All Shows button is enabled"
    print(allShowsEnabled)
    buttonEnabled.append(allShowsEnabled)
else:
    allShowsDisabled = "All shows button is disabled"
    print(allShowsDisabled)
    buttonDisabled.append(allShowsDisabled)

if documentary.is_displayed():
    documentaryDisplayed = "Documentary button is displayed"
    print(documentaryDisplayed)
    buttonDisplayed.append(documentaryDisplayed)
else:
    documentaryNotDisplayed = "Documentary button is not displayed"
    print(documentaryNotDisplayed)
    buttonNotDisplayed.append(documentaryNotDisplayed)

if documentary.is_enabled():
    documentaryEnabled = "Documentary button is enabled"
    print(documentaryEnabled )
    buttonEnabled.append(documentaryEnabled)
else:
    documentaryDisabled = "Documentary button is disabled"
    print(documentaryDisabled)
    buttonDisabled.append(documentaryDisabled)

if news.is_displayed():
    newsDisplayed = "News button is displayed"
    print(newsDisplayed)
    buttonDisplayed.append(newsDisplayed)
else:
    newsNotDisplayed = "News button is not displayed"
    print(newsNotDisplayed)
    buttonNotDisplayed.append(newsNotDisplayed)

if news.is_enabled():
    newsEnabled = "News button is enabled"
    print(newsEnabled)
    buttonEnabled.append(newsEnabled)
else:
    newsDisabled = "News button is disabled"
    print(newsDisabled)
    buttonDisabled.append(newsDisabled)

if sports.is_displayed():
    sportsDisplayed = "Sports button is displayed"
    print(sportsDisplayed)
    buttonDisplayed.append(sportsDisplayed)
else:
    sportsNotDisplayed = "Sports button is not displayed"
    print(sportsNotDisplayed)
    buttonNotDisplayed.append(sportsNotDisplayed)

if sports.is_enabled():
    sportsEnabled = "Sports button is enabled"
    print(sportsEnabled)
    buttonEnabled.append(sportsEnabled)
else:
    sportsDisabled = "Sports button is disabled"
    print(sportsDisabled)
    buttonDisabled.append(sportsDisabled)

if weather.is_displayed():
    weatherDisplayed = "Weather button is displayed"
    print(weatherDisplayed)
    buttonDisplayed.append(weatherDisplayed)
else:
    weatherNotDisplayed = "Weather button is not displayed"
    print(weatherNotDisplayed)
    buttonNotDisplayed.append(weatherNotDisplayed)

if weather.is_enabled():
    weatherEnabled = "Weather button is enabled"
    print(weatherEnabled)
    buttonEnabled.append(weatherEnabled)
else:
    weatherDisabled = "Weather button is disabled"
    print(weatherDisabled)
    buttonDisabled.append(weatherDisabled)

if nature.is_displayed():
    natureDisplayed = "Nature button is displayed"
    print(natureDisplayed)
    buttonDisplayed.append(natureDisplayed)
else:
    natureNotDisplayed = "Nature button is not displayed"
    print(natureNotDisplayed)
    buttonNotDisplayed.append(natureNotDisplayed)

if nature.is_enabled():
    natureEnabled = "Nature button is enabled"
    print(natureEnabled)
    buttonEnabled.append(natureEnabled)
else:
    natureDisabled = "Nature button is disabled"
    print(natureDisabled)
    buttonDisabled.append(natureDisabled)

if tvShows.is_displayed():
    tvShowsDisplayed = "TV Shows button is displayed"
    print(tvShowsDisplayed)
    buttonDisplayed.append(tvShowsDisplayed)
else:
    tvShowsNotDisplayed = "TV Shows button is not displayed"
    print(tvShowsNotDisplayed)
    buttonNotDisplayed.append(tvShowsNotDisplayed)

if tvShows.is_enabled():
    tvShowsEnabled = "TV Shows button is enabled"
    print(tvShowsEnabled)
    buttonEnabled.append(tvShowsEnabled)
else:
    tvShowsDisabled = "TV Shows button is disabled"
    print(tvShowsDisabled)
    buttonDisabled.append(tvShowsDisabled)

allShows.click()
returnToHomeandCat()
allShowsClickAction = "All shows selection passed"
print(allShowsClickAction)
clickActionPassed.append(allShowsClickAction)

documentary.click()
returnToHomeandCat()
documentaryClickAction = "Documentary selection passed"
print(documentaryClickAction)
clickActionPassed.append(documentaryClickAction)

news.click()
returnToHomeandCat()
newsClickAction = "News selection passed"
print(newsClickAction)
clickActionPassed.append(newsClickAction)

sports.click()
returnToHomeandCat()
sportsClickAction = "Sports selection passed"
print(sportsClickAction)
clickActionPassed.append(sportsClickAction)

weather.click()
returnToHomeandCat()
weatherClickAction = "Weather selection passed"
print("Weather selection passed")
clickActionPassed.append(weatherClickAction)

nature.click()
returnToHomeandCat()
natureClickAction = "Nature selection passed"
print(natureClickAction)
clickActionPassed.append(natureClickAction)

tvShows.click()
returnToHomeandCat()
tvShowsClickAction = "TV Shows selection passed"
print(tvShowsClickAction)
clickActionPassed.append(tvShowsClickAction)

randomClick()

time.sleep(3)

news = driver.find_element(by=By.LINK_TEXT, value="News")
news.click()
print("News search passed")

returnToHome()
time.sleep(3)

michaelJordan = driver.find_element(by=By.LINK_TEXT, value='"Michael Jordan"')
michaelJordan.click()
print("Michael Jordan search passed")

returnToHome()
time.sleep(3)

ukraine = driver.find_element(by=By.LINK_TEXT, value="Ukraine")
ukraine.click()
print("Ukraine search passed")

returnToHome()
time.sleep(3)

searchBar = driver.find_element(by=By.CLASS_NAME, value="elastic-search-input")

newman_command_Search_bar_background_video = f"{newman_path} run \"C:\Eon\JSON File\LandingPage.postman_collection.json\" --folder Search_bar_background_video"
run_newman_and_save_output(newman_command_Search_bar_background_video, "Search_bar_background_video", output_file, error_file)
run_newman_for_folder('Search_bar_background_video')

if searchBar.is_displayed():
    searchBarDisplayed = "Search Bar is displayed"
    print(searchBarDisplayed)
    buttonDisplayed.append(searchBarDisplayed)
else:
    searchBarNotDisplayed = "Search Bar is not displayed"
    print(searchBarNotDisplayed)
    buttonNotDisplayed.append(searchBarNotDisplayed)

if searchBar.is_enabled():
    searchBarEnabled = "Search Bar is enabled"
    print(searchBarEnabled)
    buttonEnabled.append(searchBarEnabled)
else:
    searchBarDisabled = "Search Bar is disabled"
    print(searchBarDisabled)
    buttonDisabled = searchBarDisabled

searchBar.click()
searchBarClickAction = "Select search bar passed"
print("Select search bar passed")
clickActionPassed.append(searchBarClickAction)



searchBar.send_keys("Pepsi")

newman_command_search_function= f"{newman_path} run \"C:\Eon\JSON File\LandingPage.postman_collection.json\" --folder search_function"
run_newman_and_save_output(newman_command_search_function, "search_function", output_file, error_file)
run_newman_for_folder('search_function')

searchBar.send_keys(Keys.RETURN)
searchResultClickAction = "Search result passed"
print(searchResultClickAction)
clickActionPassed.append(searchResultClickAction)



returnToHome()

time.sleep(4)

video = driver.find_element(by=By.CLASS_NAME, value="slider-item-thumbnail-container")
action = ActionChains(driver)

newman_command_carousels = f"{newman_path} run \"C:\Eon\JSON File\LandingPage.postman_collection.json\" --folder carousels"
run_newman_and_save_output(newman_command_carousels, "carousels", output_file, error_file)
run_newman_for_folder('carousels')

action.move_to_element(video).perform()
hoverClickAction = "Hover over thumbnail passed"
print(hoverClickAction)
hoverActionPassed.append(hoverClickAction)


time.sleep(2)

mute = driver.find_element(by=By.CLASS_NAME, value="gvpi-custom-mute-button-mini-hover")
if mute.is_displayed():
    muteDisplayed = "Mute button is displayed"
    print(muteDisplayed)
    buttonDisplayed.append(muteDisplayed)
else:
    muteNotDisplayed = "Mute button is not displayed"
    print(muteNotDisplayed)
    buttonNotDisplayed.append(muteNotDisplayed)

if mute.is_enabled():
    muteEnabled = "Mute button is enabled"
    print(muteEnabled)
    buttonEnabled.append(muteEnabled)
else:
    muteDisabled = "Mute button is disabled"
    print(muteDisabled)
    buttonDisabled.append(muteDisabled)

mute.click()
muteClickAction = "Mute passed"
print(muteClickAction)
clickActionPassed.append(muteClickAction)

time.sleep(2)
mute.click()
unmuteClickAction = "Un-mute passed"
print(unmuteClickAction)
clickActionPassed.append(unmuteClickAction)

time.sleep(1)

favorite = driver.find_element(by=By.CLASS_NAME, value="svg-icon-container")

newman_command_add_favorite= f"{newman_path} run \"C:\Eon\JSON File\LandingPage.postman_collection.json\" --folder add_favorite"
run_newman_and_save_output(newman_command_add_favorite, "add_favorite", output_file, error_file)
run_newman_for_folder('add_favorite')

if favorite.is_displayed():
    favoriteDisplayed = "Favorites button is displayed"
    print(favoriteDisplayed)
    buttonDisplayed.append(favoriteDisplayed)
else:
    favoriteNotDisplayed = "Favorites button is displayed"
    print("Favorites button is not displayed")
    buttonNotDisplayed.append(favoriteNotDisplayed)

if favorite.is_enabled():
    favoriteEnabled = "Favorites button is enabled"
    print(favoriteEnabled)
    buttonEnabled.append(favoriteEnabled)
else:
    favoriteDisabled = "Favorites button is disabled"
    print(favoriteDisabled)
    buttonDisabled.append(favoriteDisabled)

copy = driver.find_element(by=By.CLASS_NAME, value="fill-icon")
if copy.is_displayed():
    copyDisplayed = "Copy button is displayed"
    print(copyDisplayed)
    buttonDisplayed.append(copyDisplayed)
else:
    copyNotDisplayed = "Copy button is not displayed"
    print(copyNotDisplayed)
    buttonNotDisplayed.append(copyNotDisplayed)

if copy.is_enabled():
    copyEnabled = "Copy button is enabled"
    print(copyEnabled)
    buttonEnabled.append(copyEnabled)
else:
    copyDisabled = "Copy button is disabled"
    print(copyDisabled)
    buttonDisabled.append(copyDisabled)

favorite.click()
favoriteClickAction = "Favorite button passed"
print(favoriteClickAction)
clickActionPassed.append(favoriteClickAction)

time.sleep(1)

favorite.click()
unfavoriteClickAction = "Un-favorite button passed"
print(unfavoriteClickAction)
clickActionPassed.append(unfavoriteClickAction)


newman_command_remove_favourite= f"{newman_path} run \"C:\Eon\JSON File\LandingPage.postman_collection.json\" --folder remove_favourite"
run_newman_and_save_output(newman_command_remove_favourite, "remove_favourite", output_file, error_file)
run_newman_for_folder('remove_favourite')

time.sleep(2)

moreInformation = driver.find_element(by=By.CLASS_NAME, value="slider-item-icon-play")
if moreInformation.is_displayed():
    moreInformationDisplayed = "More Information button is displayed"
    print(moreInformationDisplayed)
    buttonDisplayed.append(moreInformationDisplayed)
else:
    moreInformationNotDisplayed = "More Information button is not displayed"
    print(moreInformationNotDisplayed)
    buttonNotDisplayed.append(moreInformationNotDisplayed)

if moreInformation.is_enabled():
    moreInformationEnabled = "More Information button is enabled"
    print(moreInformationEnabled)
    buttonEnabled.append(moreInformationEnabled)
else:
    moreInformationDisabled = "More Information button is disabled"
    print(moreInformationDisabled)
    buttonDisabled.append(moreInformationDisabled)

moreInformation.click()
moreInformationClickAction = "More information button passed"
print(moreInformationClickAction)
clickActionPassed.append(moreInformationClickAction)

time.sleep(2)

allDetails = driver.find_element(by=By.CLASS_NAME, value="font-vertical-fix-extra")
if allDetails.is_displayed():
    allDetailsDisplayed = "All details button is displayed"
    print(allDetailsDisplayed)
    buttonDisplayed.append(allDetailsDisplayed)
else:
    allDetailsNotDisplayed = "All details button is not displayed"
    print(allDetailsNotDisplayed)
    buttonNotDisplayed.append(allDetailsNotDisplayed)

if allDetails.is_enabled():
    allDetailsEnabled = "All details button is enabled"
    print(allDetailsEnabled)
    buttonEnabled.append(allDetailsEnabled)
else:
    allDetailsDisabled = "All details button is disabled"
    print(allDetailsDisabled)
    buttonDisabled.append(allDetailsDisabled)

allDetails.click()
allDetailsClickAction = "All details button passed"
print(allDetailsClickAction)
clickActionPassed.append(allDetailsClickAction)

returnToHome()
print("Return to home passed")

time.sleep(3)

profileIcon()

newman_command_logout= f"{newman_path} run \"C:\Eon\JSON File\LandingPage.postman_collection.json\" --folder logout"
run_newman_and_save_output(newman_command_logout, "logout", output_file, error_file)
run_newman_for_folder('logout')

logout.click()
logoutClickAction = "logout passed"
print(logoutClickAction)
clickActionPassed.append(logoutClickAction)

notDisplayed = "\n".join(buttonNotDisplayed)
disabled = "\n".join(buttonDisabled)

with open(error_file, "a", encoding="utf-8") as file:
    file.write(notDisplayed + "\n")

with open(error_file, "a", encoding="utf-8") as file:
    file.write(disabled + "\n")

ws['A1'] = "Click Actions Passed"
ws['B1'] = "Buttons Displayed"
ws['C1'] = "Buttons Enabled"
ws['D1'] = "Hover action passed"
ws['E1'] = "Buttons Not Displayed"
ws['F1'] = "Buttons Disabled"

for index, value in enumerate(clickActionPassed):
    cell = ws.cell(row=index+2, column=1, value=value)

for index, value in enumerate(buttonDisplayed):
    cell = ws.cell(row=index+2, column=2, value=value)

for index, value in enumerate(buttonEnabled):
    cell = ws.cell(row=index+2, column=3, value=value)

for index, value in enumerate(hoverActionPassed):
    cell = ws.cell(row=index+2, column=4, value=value)

for index, value in enumerate(buttonNotDisplayed):
    cell = ws.cell(row=index+2, column=5, value=value)

for index, value in enumerate(buttonDisabled):
    cell = ws.cell(row=index+2, column=6, value=value)

wb.save('Automation Test Results.xlsx')
